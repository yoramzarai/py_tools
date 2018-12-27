#!/opt/local/bin/python3
'''Displays virus or host UR results from the Viruses/VUR/ folders.'''
import argparse
import openpyxl as opxl
import re
import sys
from termcolor import colored
from pprint import pprint

# Command-line parsing information
def parse_in_args():
    ''' Defines input arguments '''
    # place description
    parser = argparse.ArgumentParser(description='Displays virus or host UR results.', add_help=False)

    # required arguments
    group = parser.add_argument_group('Required arguments')
    group.add_argument('-f', dest='file', help='VUR file to process', type=str, metavar='<file>', required='True')

    # optional arguments
    group_opt = parser.add_argument_group('Optional arguments')
    group_opt.add_argument('-t', dest='taxid', help='Display UR of this taxid. Must be present for viruses UR', \
                       type=int, metavar='<taxid>')
    group_opt.add_argument('-n', dest='hname', \
                           help='Display UR of this host name. Used ONLY for hosts and when taxid is not given.', \
                           type=str, metavar='<name>')

    group_opt.add_argument('-D', help='Enables debug prints.', action='store_true')
    group_opt.add_argument('-h', '--help', help='Displays usage information and exits.', action='help')

    #return parser.parse_args()
    #return parser
    return parser.parse_args(), parser.print_help
    
# Old version (not used) !!!!!
def process_virus_vur_old(sheet, taxid, lufld, taxids):
    '''Displays requested virus VUR entries'''
    rng = range(0,len(lufld))
    # rows with required taxid infromation
    row = [ind for ind, val in enumerate(taxids) if val.value == taxid]
    if row:
        # header information
        print(colored('{0:<10}\t', 'red').format('taxid'), end='')
        for c in rng:
            print(colored('{}\t', 'red').format(lufld[c][0].value), end='')
        print('')
        
        # data
        for cnt in row:
            print('{0:<10}\t'.format(taxid), end='')
            for c in rng:
                print('{}\t'.format(lufld[c][cnt].value),end='')
        print('')
    else:
        print('taxid = {taxid} not found!!'.format(taxid=taxid))

# Old version (not used) !!!!
def process_host_vur_old(sheet, taxid, hst_name, lufld, taxids, hst_names):
    '''Displays requested host VUR entries'''
    rng = range(0,len(lufld))
    # rows with required taxid infromation
    if taxid:
        ary = taxids
        val = taxid
    else:
        ary = hst_names
        val = hst_name
    row = [ind for ind, v in enumerate(ary) if v.value == val]
    
    if row:
        # header information
        print(colored('{0:<10}{1:<50}{2:<20}{3:<10}\t', 'red').format('taxid','name','DB','vtaxid'), end='')
        for c in rng:
            print(colored('{0}\t', 'red').format(lufld[c][0].value), end='')
        print('')
        
        # data
        for cnt in row:
            db = sheet.cell(row=cnt+1,column=1)
            name = sheet.cell(row=cnt+1,column=2)
            htaxid = sheet.cell(row=cnt+1,column=3)
            vtaxid = sheet.cell(row=cnt+1,column=4)
            print('{htaxid:<10}{name:<50}{DB:<20}{vtaxid:<10}\t'.format(htaxid=htaxid.value, name=name.value, \
                                                             DB=db.value, vtaxid=vtaxid.value), end='')
            for c in rng:
                print('{0}\t'.format(lufld[c][cnt].value),end='')
            print('')
    else:
        print('taxid = {taxid} not found!!'.format(taxid=taxid))

def process_host_vur(sheet, taxid, hst_name, lufld, taxids, hst_names):
    '''Displays requested host VUR entries'''
    rng = range(0,len(lufld))
    if taxid:
        ary = taxids
        val = taxid
    else:
        ary = hst_names
        val = hst_name
    row = [ind for ind, v in enumerate(ary) if v.value == val]  # rows with required taxid infromation
    info = {} # (htaxid, vtaxid) as key, list of requested fields as value
    if row:
        # header information
        print(colored('{0:<10}{1:<10}\t', 'red').format('htaxid','vtaxid'), end='')
        for c in rng: print(colored('{0}\t\t', 'red').format(lufld[c][0].value), end='')
        print('')
        
        # data
        for cnt in row:
            htaxid = sheet.cell(row=cnt+1,column=3).value
            vtaxid = sheet.cell(row=cnt+1,column=4).value
            info[(htaxid,vtaxid)]=[lufld[c][cnt].value for c in rng]
            print('{htaxid:<10}{vtaxid:<10}'.format(htaxid=htaxid, vtaxid=vtaxid), end='\t')
            print(*info[(htaxid,vtaxid)], sep='\t\t')
    else:
        print('taxid = {taxid} not found!!'.format(taxid=taxid))
    return info
        
def process_virus_vur(sheet, taxid, lufld, taxids):
    '''Displays requested virus VUR entries'''
    rng = range(0,len(lufld))
    info = {} # taxid as key, list of requested fields as value
    row = [ind for ind, val in enumerate(taxids) if val.value == taxid] 
    if row:
        row = row.pop() # only one in case of viruses VUR
        # header information
        print(colored('{0:<10}\t', 'red').format('taxid'), end='')
        for c in rng: print(colored('{}\t', 'red').format(lufld[c][0].value), end='')
        print('')
        
        # data
        info[taxid]=[lufld[c][row].value for c in rng]
        print('{taxid:<10}'.format(taxid=taxid), end='\t')
        print(*info[taxid], sep='\t')
    else:
        print('taxid = {taxid} not found!!'.format(taxid=taxid))
    return info

def check_params_valid(vtype, args, print_help):
    '''Checking that we have the necessary inpput arguments'''
    if vtype == 'viruses':
        if args.taxid == None:  # the -t option is required
            print('Error: -t option is required for the viruses VUR file !!')
            print_help()
            sys.exit('Exit: Missing input arguments')
    else:
        if args.taxid == None and args.hname == None:  # either -t or -n is required
            print('Error: -t or -n option is required for the hosts VUR file !!')
            print_help()
            sys.exit('Exit: Missing input arguments')
    return True

# Main function
# =============
def main():
    #args = parse_in_args()  # parse, validate and return input arguments
    #parser = parse_in_args()  # parse, validate and return input arguments
    #args = parser.parse_args()
    args, parser_print_help = parse_in_args()  # parse, validate and return input arguments
    
    xd = opxl.load_workbook(args.file)
    sheet = xd.active
    if sheet.cell(row=1,column=1).value == 'vtaxid': # if A(1,1)='vtaxid' then this is a virus VUR
        taxid_str='vtaxid'
        hst_name_str = ''
        vtype = 'viruses'
    else:
        taxid_str = 'hst_taxid'
        hst_name_str = 'hst_name'
        vtype = 'hosts'        
    if args.D: print('Processing {type} VUR...'.format(type=vtype))

    # check that we have the necessary arguments
    check_params_valid(vtype, args, parser_print_help)

    dm = re.split(':',sheet.dimensions)
    if args.D: print('xlsx dimension= {}:{}'.format(dm[0],dm[1]))
    
    ab = list(map(chr, range(65, 91))) # create uppercase alphabet
    print('{} rows found and the following fields (columns):'.format(int(dm[1][1:])-1))
    for cnt in range(1,sheet.max_column+1):
        cobj = sheet.cell(row=1,column=cnt)
        print('{indx}. {field}'.format(indx=ab[cnt-1], field=cobj.value))
        if re.search(taxid_str, cobj.value): # this column contains taxid 
            taxids_sheet_idx = ab[cnt-1]
        if re.search(hst_name_str, cobj.value): # this column contains host name
            hst_name_sheet_idx = ab[cnt-1]

    fidx = input("Which fields to display (if multiple, use space in between)? ")
    print('')
    lufld = list(map(lambda x: sheet[x], fidx.split(' ')))
    taxids = sheet[ taxids_sheet_idx ]
    if vtype=='hosts': hst_names = sheet[ hst_name_sheet_idx ]

    # process the VUR
    if vtype=='viruses':
        info = process_virus_vur(sheet, args.taxid, lufld, taxids)
    else:
        info = process_host_vur(sheet, args.taxid, args.hname, lufld, taxids, hst_names)

    pprint( info )
# =============================================================================================================================
if __name__ == "__main__":
    main()

